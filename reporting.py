import os
import cv2
import shutil
import numpy as np
from datetime import datetime
import csv
import pandas as pd
from pathlib import Path
import config
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, Alignment, PatternFill

class Reporter:
    def __init__(self, evidence_path, report_path):
        """Initialize with paths for evidence and reports."""
        self.evidence_path = Path(evidence_path)
        self.report_path = Path(report_path)
        self.evidence_path.mkdir(parents=True, exist_ok=True)
        self.report_path.mkdir(parents=True, exist_ok=True)

    def save_evidence(self, event):
        """Save video clip and images using event's frame_count."""
        timestamp_str = event['timestamp'].strftime('%Y%m%d_%H%M%S')
        folder = self.evidence_path / f"event_{timestamp_str}_{event['vehicle_id']}"
        folder.mkdir(parents=True, exist_ok=True)

        video_path = getattr(config, 'video_path', None)
        if not video_path or not os.path.exists(video_path):
            print(f"Error: Invalid video_path in config: {video_path}")
            return None, None

        cap = cv2.VideoCapture(video_path)
        if not cap.isOpened():
            print(f"Error: Could not open video {video_path}")
            return None, None

        fps = cap.get(cv2.CAP_PROP_FPS)
        total_frames = int(cap.get(cv2.CAP_PROP_FRAME_COUNT))
        start_frame = max(0, min(event['frame_count'] - int(5 * fps), total_frames - 1))
        cap.set(cv2.CAP_PROP_POS_FRAMES, start_frame)

        fourcc = cv2.VideoWriter_fourcc(*'avc1')
        width = int(cap.get(cv2.CAP_PROP_FRAME_WIDTH))
        height = int(cap.get(cv2.CAP_PROP_FRAME_HEIGHT))
        clip_path = str(folder / 'clip.mp4')
        out = cv2.VideoWriter(clip_path, fourcc, fps, (width, height))

        frames_to_capture = int(10 * fps)
        middle_frame = int(frames_to_capture / 2)
        before_img_path = str(folder / 'before.jpg')
        after_img_path = str(folder / 'after.jpg')

        for i in range(frames_to_capture):
            ret, frame = cap.read()
            if not ret:
                break
            if i == 0:
                cv2.imwrite(before_img_path, frame)
            if i == middle_frame:
                cv2.imwrite(after_img_path, frame)
            out.write(frame)

        out.release()
        cap.release()
        
        if os.path.getsize(clip_path) > 0:
            return before_img_path, after_img_path
        else:
            print("Warning: Evidence video clip is empty")
            return before_img_path, after_img_path

    def export_events(self, events_data, camera_id):
        """Export events to CSV and Excel with evidence."""
        if not events_data:
            print("No events to report")
            return None

        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        report_base = f"report_{timestamp}_{camera_id}"
        csv_file = self.report_path / f"{report_base}.csv"
        excel_file = self.report_path / f"{report_base}.xlsx"

        wb = Workbook()
        ws = wb.active
        ws.title = "Waste Disposal Events"

        headers = ['Timestamp', 'Camera ID', 'Vehicle ID', 'Event Type', 'Location', 
                  'Velocity', 'Confidence', 'Before Image', 'After Image', 'Evidence Path']
        header_fill = PatternFill(start_color="66CCFF", end_color="66CCFF", fill_type="solid")
        header_font = Font(bold=True, size=12)

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            ws.column_dimensions[chr(64 + col)].width = 18
        ws.column_dimensions['H'].width = 30
        ws.column_dimensions['I'].width = 30

        csv_data = []
        excel_row = 2

        with open(csv_file, 'w', newline='') as f:
            writer = csv.writer(f)
            writer.writerow(headers)

            for event in events_data:
                before_img_path, after_img_path = self.save_evidence(event)
                
                event_data = [
                    event['timestamp'].strftime('%Y-%m-%d %H:%M:%S'),
                    camera_id,
                    event['vehicle_id'],
                    event['event_type'],
                    str(event['location']),
                    f"{event['velocity']:.2f}",
                    f"{event.get('confidence', 0.0):.2f}"
                ]

                csv_row = event_data.copy()
                evidence_path = before_img_path.replace('before.jpg', 'clip.mp4') if before_img_path else 'N/A'
                csv_row.extend([before_img_path or 'N/A', after_img_path or 'N/A', evidence_path])
                writer.writerow(csv_row)
                csv_data.append(csv_row)

                for col, value in enumerate(event_data, 1):
                    ws.cell(row=excel_row, column=col, value=value)

                if before_img_path and os.path.exists(before_img_path):
                    try:
                        img = XLImage(before_img_path)
                        img.width = 300
                        img.height = 200
                        ws.add_image(img, f'H{excel_row}')
                    except Exception as e:
                        print(f"Failed to add before image: {e}")
                        ws.cell(row=excel_row, column=8, value="Image failed")

                if after_img_path and os.path.exists(after_img_path):
                    try:
                        img = XLImage(after_img_path)
                        img.width = 300
                        img.height = 200
                        ws.add_image(img, f'I{excel_row}')
                    except Exception as e:
                        print(f"Failed to add after image: {e}")
                        ws.cell(row=excel_row, column=9, value="Image failed")

                ws.cell(row=excel_row, column=10, value=evidence_path)
                excel_row += 1

        try:
            wb.save(excel_file)
            print(f"Excel report generated: {excel_file}")
        except Exception as e:
            print(f"Failed to save Excel report: {e}")
            return None

        return {'csv': str(csv_file), 'excel': str(excel_file), 'events': csv_data}